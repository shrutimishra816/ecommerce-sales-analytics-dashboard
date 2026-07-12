"""
Quality & Inventory Adjustment Root-Cause Report
--------------------------------------------------
Takes the cleaned transaction warehouse (produced by etl/etl_pipeline.py),
enriches every returned order with a defect/adjustment reason, and publishes
a stakeholder-ready Excel workbook:

    1. Summary        - headline KPIs (return rate, top defect reason, etc.)
    2. Root Cause      - Pareto-style ranking of defect reasons (80/20 view)
    3. By Category     - defect rate + top reason per product category
    4. By Region        - defect rate + top reason per region
    5. Adjustment Log   - row-level inventory adjustment detail for audit

This mirrors the kind of daily/weekly/monthly quality + inventory-adjustment
reporting a retail/warehouse quality analyst would publish, built entirely
in Python + Excel (openpyxl) rather than a BI tool, since Excel is the
delivery format most operations stakeholders actually work in.

Run: python excel_reporting/quality_root_cause_report.py
"""
import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

DB_PATH = Path(__file__).parent.parent / "data" / "ecommerce_warehouse.db"
OUT_PATH = Path(__file__).parent / "quality_root_cause_report.xlsx"

rng = np.random.default_rng(7)

# Defect / adjustment reason taxonomy, weighted the way real retail return
# reasons tend to skew (majority "quality/fit" issues, a long tail of
# logistics + fulfillment causes) -- this is the root-cause categorization
# layer that was missing from the original return-rate-only analysis.
DEFECT_REASONS = [
    ("Quality Defect - Item Damaged", 0.27),
    ("Wrong Item Shipped", 0.18),
    ("Size / Fit Issue", 0.16),
    ("Damaged in Transit", 0.14),
    ("Late Delivery - Order Cancelled", 0.10),
    ("Packaging Failure", 0.08),
    ("Customer Changed Mind", 0.07),
]
reason_names = [r for r, _ in DEFECT_REASONS]
reason_weights = [w for _, w in DEFECT_REASONS]

HEADER_FILL = PatternFill("solid", fgColor="232F3E")   # Amazon-navy header
ACCENT_FILL = PatternFill("solid", fgColor="FF9900")   # Amazon-orange accent
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="232F3E")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_data() -> pd.DataFrame:
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT * FROM fact_transactions", con)
    con.close()
    return df


def enrich_with_defect_reasons(df: pd.DataFrame) -> pd.DataFrame:
    returns = df[df["is_returned"] == 1].copy()
    returns["defect_reason"] = rng.choice(reason_names, size=len(returns), p=reason_weights)
    # Adjustment quantity: units pulled from sellable inventory because of the defect
    returns["adjustment_qty"] = returns["quantity"]
    returns["adjustment_value"] = returns["net_revenue"]
    returns["severity"] = np.select(
        [returns["defect_reason"].isin(["Quality Defect - Item Damaged", "Damaged in Transit"]),
         returns["defect_reason"].isin(["Wrong Item Shipped", "Packaging Failure"])],
        ["High", "Medium"],
        default="Low",
    )
    return returns


def style_header(ws, row, ncols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, ncols, width=20):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def write_df(ws, df, start_row=1, start_col=1):
    ncols = len(df.columns)
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    style_header(ws, start_row, ncols)
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=val)
            cell.border = BORDER
    return start_row + len(df) + 1


def build_report():
    df = load_data()
    returns = enrich_with_defect_reasons(df)

    total_orders = len(df)
    total_returns = len(returns)
    return_rate = total_returns / total_orders
    total_adj_value = returns["adjustment_value"].sum()

    pareto = (
        returns.groupby("defect_reason")
        .agg(orders=("order_id", "count"), adjustment_value=("adjustment_value", "sum"))
        .sort_values("orders", ascending=False)
        .reset_index()
    )
    pareto["pct_of_total"] = (pareto["orders"] / total_returns * 100).round(1)
    pareto["cumulative_pct"] = pareto["pct_of_total"].cumsum().round(1)

    by_category = (
        df.groupby("category")
        .apply(lambda g: pd.Series({
            "orders": len(g),
            "returns": g["is_returned"].sum(),
            "defect_rate_pct": round(g["is_returned"].mean() * 100, 2),
        }))
        .reset_index()
    )
    top_reason_by_cat = (
        returns.groupby("category")["defect_reason"]
        .agg(lambda s: s.value_counts().idxmax())
        .reset_index()
        .rename(columns={"defect_reason": "top_defect_reason"})
    )
    by_category = by_category.merge(top_reason_by_cat, on="category", how="left")
    by_category = by_category.sort_values("defect_rate_pct", ascending=False)

    by_region = (
        df.groupby("region")
        .apply(lambda g: pd.Series({
            "orders": len(g),
            "returns": g["is_returned"].sum(),
            "defect_rate_pct": round(g["is_returned"].mean() * 100, 2),
        }))
        .reset_index()
    )
    top_reason_by_region = (
        returns.groupby("region")["defect_reason"]
        .agg(lambda s: s.value_counts().idxmax())
        .reset_index()
        .rename(columns={"defect_reason": "top_defect_reason"})
    )
    by_region = by_region.merge(top_reason_by_region, on="region", how="left")
    by_region = by_region.sort_values("defect_rate_pct", ascending=False)

    adj_log_cols = ["order_id", "order_date", "region", "category", "sku",
                     "defect_reason", "severity", "adjustment_qty", "adjustment_value"]
    adjustment_log = returns[adj_log_cols].sort_values("order_date", ascending=False).head(500)

    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["B2"] = "E-Commerce Quality & Inventory Adjustment Report"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Root-cause breakdown of returns, defect categorization, and inventory adjustment log"
    ws["B3"].font = Font(italic=True, color="595959")

    kpis = [
        ("Total Orders", f"{total_orders:,}"),
        ("Total Returns / Adjustments", f"{total_returns:,}"),
        ("Overall Return Rate", f"{return_rate:.2%}"),
        ("Total Adjustment Value ($)", f"${total_adj_value:,.0f}"),
        ("#1 Root Cause", pareto.iloc[0]["defect_reason"]),
        ("#1 Root Cause Share of Returns", f"{pareto.iloc[0]['pct_of_total']}%"),
    ]
    r = 5
    for label, val in kpis:
        ws.cell(row=r, column=2, value=label).font = Font(bold=True)
        ws.cell(row=r, column=3, value=val)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="F2F2F2")
        r += 1
    autosize(ws, 4, width=34)

    # ---- Root Cause (Pareto) ----
    ws2 = wb.create_sheet("Root Cause (Pareto)")
    ws2["A1"] = "Return Root-Cause Pareto Analysis"
    ws2["A1"].font = TITLE_FONT
    next_row = write_df(ws2, pareto, start_row=3)
    autosize(ws2, len(pareto.columns))

    chart = BarChart()
    chart.title = "Returns by Root Cause"
    chart.y_axis.title = "Orders"
    chart.x_axis.title = "Defect Reason"
    data_ref = Reference(ws2, min_col=2, min_row=3, max_row=2 + len(pareto))
    cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(pareto))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 22
    chart.height = 10
    ws2.add_chart(chart, f"A{next_row + 1}")

    # ---- By Category ----
    ws3 = wb.create_sheet("By Category")
    ws3["A1"] = "Defect Rate by Product Category"
    ws3["A1"].font = TITLE_FONT
    write_df(ws3, by_category, start_row=3)
    autosize(ws3, len(by_category.columns))
    rule = ColorScaleRule(start_type="min", start_color="63BE7B",
                           end_type="max", end_color="F8696B")
    ws3.conditional_formatting.add(f"D4:D{3 + len(by_category)}", rule)

    # ---- By Region ----
    ws4 = wb.create_sheet("By Region")
    ws4["A1"] = "Defect Rate by Region"
    ws4["A1"].font = TITLE_FONT
    write_df(ws4, by_region, start_row=3)
    autosize(ws4, len(by_region.columns))
    ws4.conditional_formatting.add(f"D4:D{3 + len(by_region)}", rule)

    # ---- Adjustment Log ----
    ws5 = wb.create_sheet("Adjustment Log")
    ws5["A1"] = "Inventory Adjustment Log (most recent 500)"
    ws5["A1"].font = TITLE_FONT
    write_df(ws5, adjustment_log, start_row=3)
    autosize(ws5, len(adjustment_log.columns), width=18)
    sev_col = adj_log_cols.index("severity") + 1
    high_fill = PatternFill("solid", fgColor="F8696B")
    for i in range(len(adjustment_log)):
        cell = ws5.cell(row=4 + i, column=sev_col)
        if cell.value == "High":
            cell.fill = high_fill
            cell.font = Font(color="FFFFFF", bold=True)

    wb.save(OUT_PATH)
    print(f"Saved report -> {OUT_PATH}")
    print(f"Total orders: {total_orders:,} | Returns: {total_returns:,} ({return_rate:.2%})")
    print(f"Top root cause: {pareto.iloc[0]['defect_reason']} ({pareto.iloc[0]['pct_of_total']}%)")


if __name__ == "__main__":
    build_report()
